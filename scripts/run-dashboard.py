#!/usr/bin/env python3

from __future__ import annotations

import argparse
import html
import json
import os
import re
from dataclasses import dataclass
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse, quote

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
LIBS_DIR = ROOT / "libs"
OUTPUTS_DIR = ROOT / "outputs"
REPORTS_DIR = ROOT / "reports"
DASHBOARD_DIR = ROOT / "dashboard"


HEADER_ALIASES = {
    "库名": "lib_name",
    "git仓库": "repo_url",
    "版本": "version",
    "是否需要用户审批方案": "approval_required",
    "审批结果": "approval_result",
    "适配状态": "adaptation_status",
    "编译状态": "build_status",
    "测试状态": "test_status",
    "失败原因/备注": "note",
}


@dataclass
class BatchSummary:
    date: str
    path: str
    libs: list[str]
    rows: list[dict[str, str]]
    summary: list[str]


def normalize_header(text: str) -> str:
    value = (text or "").strip().lower()
    value = value.replace(" ", "")
    value = value.replace("（", "(").replace("）", ")")
    value = re.sub(r"\(.*?\)", "", value)
    return value


def find_task_files() -> list[Path]:
    return sorted(LIBS_DIR.glob("porting-tasks-????-??-??.xlsx"))


def iter_non_dir_files(path: Path) -> list[Path]:
    if not path.exists():
        return []
    return [item for item in path.rglob("*") if item.is_file()]


def rel(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except ValueError:
        return path.as_posix()


def parse_task_sheet(task_file: Path) -> list[dict[str, Any]]:
    wb = load_workbook(task_file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    raw_header = [("" if c is None else str(c).strip()) for c in rows[0]]
    header_keys = []
    for cell in raw_header:
        header_keys.append(HEADER_ALIASES.get(normalize_header(cell), ""))

    defaults = {
        "lib_name": "",
        "repo_url": "",
        "version": "",
        "approval_required": "是",
        "approval_result": "",
        "adaptation_status": "待处理",
        "build_status": "待处理",
        "test_status": "待处理",
        "note": "",
    }

    result = []
    for index, row in enumerate(rows[1:], start=2):
        values = [("" if c is None else str(c).strip()) for c in row]
        item = dict(defaults)
        for idx, key in enumerate(header_keys):
            if key and idx < len(values):
                item[key] = values[idx]

        if not item["lib_name"] or not item["repo_url"]:
            continue

        if not item["approval_required"]:
            item["approval_required"] = "是"

        if item["approval_required"] == "否" and not item["approval_result"]:
            item["approval_result"] = "不需要审批"
        elif item["approval_required"] != "否" and not item["approval_result"]:
            item["approval_result"] = "待审批"

        item["_sheet_row"] = index
        date_match = re.search(r"(\d{4}-\d{2}-\d{2})", task_file.name)
        item["_task_date"] = date_match.group(1) if date_match else ""
        item["_task_file"] = rel(task_file)
        result.append(item)

    result.sort(key=lambda x: (0 if x["approval_required"] == "否" else 1, x["_sheet_row"]))
    return result


def parse_batch_report(path: Path) -> BatchSummary:
    text = path.read_text(encoding="utf-8", errors="ignore")
    date_match = re.search(r"批次汇总报告（(\d{4}-\d{2}-\d{2})）", text)
    date = date_match.group(1) if date_match else path.stem.removeprefix("batch-")

    libs = []
    libs_match = re.search(r"本轮批次库：(.*)", text)
    if libs_match:
        libs = [part.strip("` ") for part in libs_match.group(1).split("、") if part.strip()]

    rows: list[dict[str, str]] = []
    table_match = re.search(
        r"\| 库名 \| 是否需要用户审批方案 \| 审批结果 \| 适配状态 \| 编译状态 \| 测试状态 \| 失败原因/备注 \|\n"
        r"\|[-| ]+\|\n"
        r"(?P<body>(?:\|.*\|\n?)*)",
        text,
    )
    if table_match:
        body = table_match.group("body")
        for line in body.splitlines():
            if not line.startswith("|"):
                break
            parts = [part.strip() for part in line.strip().strip("|").split("|")]
            if len(parts) != 7 or parts[0] == "库名":
                continue
            rows.append(
                {
                    "lib_name": parts[0],
                    "approval_required": parts[1],
                    "approval_result": parts[2],
                    "adaptation_status": parts[3],
                    "build_status": parts[4],
                    "test_status": parts[5],
                    "note": parts[6],
                }
            )

    summary: list[str] = []
    section_match = re.search(r"## 4\. 交付总结\s+(?P<body>.*)", text, re.S)
    if section_match:
        for line in section_match.group("body").splitlines():
            line = line.strip()
            if line.startswith("- "):
                summary.append(line[2:].strip())

    return BatchSummary(date=date, path=rel(path), libs=libs, rows=rows, summary=summary)


def build_row_state(row: dict[str, Any], task_date: str) -> dict[str, Any]:
    lib_name = row["lib_name"]
    report_dir = REPORTS_DIR / lib_name
    output_dir = OUTPUTS_DIR / lib_name
    source_dir = LIBS_DIR / lib_name

    adaptation_plan = report_dir / "adaptation-plan.md"
    adaptation_report = report_dir / "adaptation-report.md"
    build_report = report_dir / "build-report.md"
    lib_dir = output_dir / "lib"
    bin_dir = output_dir / "bin"

    lib_files = iter_non_dir_files(lib_dir)
    bin_files = iter_non_dir_files(bin_dir)
    has_so = any(".so" in item.name for item in lib_files)
    has_binary = len(bin_files) > 0

    anomalies: list[str] = []
    if row["build_status"] == "pass" and not has_so:
        anomalies.append("任务表显示编译通过，但未发现 .so 产物")
    if row["test_status"] == "pass" and not has_binary:
        anomalies.append("任务表显示测试通过，但未发现 binary 产物")
    if row["adaptation_status"] == "pass" and not adaptation_report.exists():
        anomalies.append("任务表显示适配通过，但缺少 adaptation-report.md")
    if row["build_status"] == "pass" and not build_report.exists():
        anomalies.append("任务表显示编译通过，但缺少 build-report.md")

    approval_required = row["approval_required"] or "是"
    approval_result = row["approval_result"] or ("不需要审批" if approval_required == "否" else "待审批")
    adaptation_status = row["adaptation_status"] or "待处理"
    build_status = row["build_status"] or "待处理"
    test_status = row["test_status"] or "待处理"

    if anomalies:
        derived = "异常"
    elif approval_required != "否" and approval_result == "待审批":
        derived = "方案待审批"
    elif "fail" in {adaptation_status, build_status, test_status}:
        derived = "失败"
    elif adaptation_status == "pass" and build_status == "pass" and test_status == "pass":
        derived = "已完成"
    elif build_status == "pass" and test_status in {"待处理", "skip", ""}:
        derived = "已编译未测试"
    elif adaptation_status == "pass" and build_status in {"待处理", "", "skip"}:
        derived = "已适配"
    elif adaptation_plan.exists() or adaptation_report.exists():
        derived = "适配中"
    elif source_dir.exists():
        derived = "源码已拉取"
    else:
        derived = "待处理"

    return {
        "lib_name": lib_name,
        "version": row["version"],
        "task_date": task_date,
        "approval_required": approval_required,
        "approval_result": approval_result,
        "adaptation_status": adaptation_status,
        "build_status": build_status,
        "test_status": test_status,
        "note": row["note"],
        "derived_status": derived,
        "anomalies": anomalies,
        "paths": {
            "source_dir": rel(source_dir),
            "report_dir": rel(report_dir),
            "output_dir": rel(output_dir),
            "adaptation_plan": rel(adaptation_plan) if adaptation_plan.exists() else "",
            "adaptation_report": rel(adaptation_report) if adaptation_report.exists() else "",
            "build_report": rel(build_report) if build_report.exists() else "",
        },
        "artifacts": {
            "has_source_dir": source_dir.exists(),
            "has_report_dir": report_dir.exists(),
            "has_build_report": build_report.exists(),
            "has_output_dir": output_dir.exists(),
            "has_so": has_so,
            "has_binary": has_binary,
            "lib_files": [rel(item) for item in lib_files[:20]],
            "bin_files": [rel(item) for item in bin_files[:20]],
        },
    }


def collect_status() -> dict[str, Any]:
    task_files = find_task_files()
    latest_task_file = task_files[-1] if task_files else None
    latest_date = ""
    if latest_task_file:
        match = re.search(r"(\d{4}-\d{2}-\d{2})", latest_task_file.name)
        latest_date = match.group(1) if match else latest_task_file.name

    merged_rows: dict[str, dict[str, Any]] = {}
    for task_file in task_files:
        for row in parse_task_sheet(task_file):
            key = row["lib_name"]
            current = merged_rows.get(key)
            if current is None:
                merged_rows[key] = row
                continue
            current_key = (current.get("_task_date", ""), current.get("_sheet_row", 0))
            next_key = (row.get("_task_date", ""), row.get("_sheet_row", 0))
            if next_key >= current_key:
                merged_rows[key] = row

    current_items = [
        build_row_state(row, row.get("_task_date", "")) for row in merged_rows.values()
    ]

    summary = {
        "total": len(current_items),
        "done": 0,
        "pending_approval": 0,
        "adapting": 0,
        "adapted": 0,
        "built_not_tested": 0,
        "failed": 0,
        "anomaly": 0,
    }

    for item in current_items:
        state = item["derived_status"]
        if state == "已完成":
            summary["done"] += 1
        elif state == "方案待审批":
            summary["pending_approval"] += 1
        elif state == "适配中":
            summary["adapting"] += 1
        elif state == "已适配":
            summary["adapted"] += 1
        elif state == "已编译未测试":
            summary["built_not_tested"] += 1
        elif state == "失败":
            summary["failed"] += 1
        elif state == "异常":
            summary["anomaly"] += 1

    def sort_key(item: dict[str, Any]) -> tuple[int, str, str]:
        completed = 1 if item["derived_status"] == "已完成" else 0
        return (completed, item.get("task_date", ""), item["lib_name"].lower())

    current_items.sort(key=sort_key, reverse=False)
    current_items.sort(key=lambda x: x.get("task_date", ""), reverse=True)
    current_items.sort(key=lambda x: 1 if x["derived_status"] == "已完成" else 0)

    batches = []
    for batch_file in sorted(REPORTS_DIR.glob("batch-????-??-??.md"), reverse=True):
        batches.append(parse_batch_report(batch_file))

    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "task_files": [rel(path) for path in reversed(task_files)],
        "latest_task_file": rel(latest_task_file) if latest_task_file else "",
        "latest_task_date": latest_date,
        "summary": summary,
        "current_items": current_items,
        "history": [
            {
                "date": batch.date,
                "path": batch.path,
                "libs": batch.libs,
                "rows": batch.rows,
                "summary": batch.summary,
            }
            for batch in batches
        ],
    }


def render_directory(path: Path) -> str:
    items = sorted(path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
    lines = [f"<h2>{html.escape(rel(path))}</h2>", "<ul>"]
    if path != ROOT:
        parent = path.parent
        lines.append(
            f'<li><a href="/browse?path={quote(rel(parent))}">..</a></li>'
        )
    for item in items:
        target = quote(rel(item))
        label = html.escape(item.name + ("/" if item.is_dir() else ""))
        lines.append(f'<li><a href="/browse?path={target}">{label}</a></li>')
    lines.append("</ul>")
    return "\n".join(lines)


def render_file(path: Path) -> str:
    content = path.read_text(encoding="utf-8", errors="ignore")
    return (
        f"<h2>{html.escape(rel(path))}</h2>"
        f'<p><a href="/browse?path={quote(rel(path.parent))}">返回上级</a></p>'
        f"<pre>{html.escape(content)}</pre>"
    )


class DashboardHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args: Any, **kwargs: Any) -> None:
        super().__init__(*args, directory=str(DASHBOARD_DIR), **kwargs)

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path == "/api/status":
            payload = json.dumps(collect_status(), ensure_ascii=False).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        if parsed.path == "/browse":
            query = parse_qs(parsed.query)
            raw_path = query.get("path", [""])[0]
            target = (ROOT / raw_path).resolve()
            if not str(target).startswith(str(ROOT.resolve())) or not target.exists():
                self.send_error(HTTPStatus.NOT_FOUND, "Path not found")
                return
            if target.is_dir():
                body = render_directory(target)
            else:
                body = render_file(target)
            payload = (
                "<!doctype html><html><head><meta charset='utf-8'>"
                "<title>Dashboard Browser</title>"
                "<style>body{font-family:Consolas,monospace;padding:24px;background:#f8f6ef;color:#222}"
                "a{color:#0b5cab;text-decoration:none}pre{white-space:pre-wrap;background:#fff;padding:16px;"
                "border:1px solid #ddd;border-radius:8px}ul{line-height:1.8}</style></head><body>"
                f"{body}</body></html>"
            ).encode("utf-8")
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Content-Length", str(len(payload)))
            self.end_headers()
            self.wfile.write(payload)
            return

        return super().do_GET()

    def log_message(self, format: str, *args: Any) -> None:  # noqa: A003
        return


def main() -> None:
    parser = argparse.ArgumentParser(description="Run local dashboard for ho-thirdparty-porting.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    args = parser.parse_args()

    server = ThreadingHTTPServer((args.host, args.port), DashboardHandler)
    print(f"Dashboard running at http://{args.host}:{args.port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
